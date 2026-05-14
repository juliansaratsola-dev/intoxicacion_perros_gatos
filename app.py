from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, send_from_directory
from flask_cors import CORS
import sqlite3
import json
import re
import os
import sys
import csv
import io
import hashlib
from datetime import datetime
from functools import wraps
import argparse
import hashlib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import psycopg2
import psycopg2.extras

# ---------------------------------------------------------------------------
# PyInstaller compatibility
# When frozen (--onefile exe), sys.executable is the .exe path and
# sys._MEIPASS is the temp dir where bundled assets are extracted.
# Mutable data (surveys, DB) live next to the exe; read-only assets
# (templates, static) are bundled inside the exe via --add-data.
# ---------------------------------------------------------------------------
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)   # directory of the .exe
    BUNDLE_DIR = sys._MEIPASS                     # bundled templates/static
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    BUNDLE_DIR = BASE_DIR

SURVEYS_DIR = os.path.join(BASE_DIR, 'surveys')
DB_PATH = os.path.join(BASE_DIR, 'instance', 'survey.db')
EXCEL_DIR = os.path.join(BASE_DIR, 'instance', 'exports')
REACT_BUILD_DIR = os.path.join(BUNDLE_DIR, 'frontend_build')

# PostgreSQL connection string — override via env var PG_DSN
PG_DSN = os.environ.get(
    'PG_DSN',
    'host=localhost dbname=survey.db user=julian.saratsola'
)

def get_pg_conn():
    return psycopg2.connect(PG_DSN)

def save_to_postgres(survey_id, data):
    try:
        conn = get_pg_conn()
        cur = conn.cursor()
        cur.execute(
            'INSERT INTO survey_submissions (survey_id, response_json) VALUES (%s, %s)',
            (survey_id, json.dumps(data, ensure_ascii=False))
        )
        conn.commit()
        cur.close()
        conn.close()
        return True
    except Exception as e:
        print(f'[PG] Error saving to postgres: {e}')
        return False

app = Flask(
    __name__,
    template_folder=os.path.join(BUNDLE_DIR, 'templates'),
    static_folder=os.path.join(BUNDLE_DIR, 'static'),
)
CORS(app)

# Admin token: env var > config.json > fallback 'changeme'
_config_path = os.path.join(BASE_DIR, 'config.json')
_config: dict = {}
if os.path.exists(_config_path):
    with open(_config_path, 'r', encoding='utf-8') as _f:
        _config = json.load(_f)
ADMIN_TOKEN = os.environ.get('ADMIN_TOKEN') or _config.get('ADMIN_TOKEN', 'changeme')

def require_admin(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        token = request.headers.get('X-Admin-Token') or request.args.get('admin_token')
        if token != ADMIN_TOKEN:
            return jsonify({'status': 'error', 'errors': ['Unauthorized']}), 401
        return f(*args, **kwargs)
    return wrapped

def init_db():
    os.makedirs(os.path.join(BASE_DIR, 'instance'), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS responses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            survey_id TEXT,
            response_json TEXT,
            submitted_at TEXT
        )
    ''')
    conn.commit()
    conn.close()

def load_survey(sid):
    path = os.path.join(SURVEYS_DIR, f"{sid}.json")
    if not os.path.exists(path):
        return None
    with open(path, 'r', encoding='utf-8') as f:
        return json.load(f)

@app.route('/')
def index():
    surveys = []
    if os.path.exists(SURVEYS_DIR):
        for fn in os.listdir(SURVEYS_DIR):
            if fn.endswith('.json'):
                sid = fn[:-5]
                with open(os.path.join(SURVEYS_DIR, fn), 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                surveys.append({'id': sid, 'title': meta.get('title', sid)})
    return render_template('index.html', surveys=surveys, button_style=True)

# Update the route for viewing a specific survey
@app.route('/surveys/<sid>')
def survey_view(sid):
    survey = load_survey(sid)
    if survey is None:
        return 'Encuesta no encontrada', 404

    description = (
        "Encuesta de intoxicaciones en Pequeños Animales\n"
        "Objetivos\n"
        "Mediante la presente encuesta se pretende recabar información sobre las principales\n"
        "intoxicaciones observadas en pequeños animales atendidos en clínicas veterinarias del\n"
        "Uruguay. La información proporcionada deberá corresponder a casos registrados en el periodo\n"
        "comprendido entre junio de 2025 y la actualidad.\n"
        "Los datos obtenidos serán utilizados en el marco de un proyecto de investigación CIDEC y de\n"
        "una tesis de grado de la Facultad de Veterinaria (UdelaR). A partir de los resultados relevados,\n"
        "se busca contribuir a la puesta a punto en el Laboratorio de Toxicología en dicha institución,\n"
        "de técnicas analíticas que permitan mejorar el diagnóstico de las intoxicaciones mencionadas.\n"
        "La encuesta está organizada en tres bloques: datos generales, principales tóxicos en\n"
        "veterinaria y consideraciones finales. Las preguntas referidas a los principales tóxicos se\n"
        "agrupan en las siguientes categorías: pesticidas, medicamentos de uso veterinario,\n"
        "medicamentos de uso humano, plantas y miscelanea.\n"
        "Completar esta encuesta le llevará solo unos minutos. Toda información proporcionada es\n"
        "estrictamente confidencial y se utilizará exclusivamente con fines académicos y de\n"
        "investigación."
    )

    return render_template('survey.html', survey=survey, sid=sid, description=description)

@app.route('/surveys/new')
def new_survey():
    # Load the new survey JSON
    survey_path = os.path.join(BASE_DIR, 'files', 'encuesta_version.json')
    if not os.path.exists(survey_path):
        return 'Encuesta no encontrada', 404

    with open(survey_path, 'r', encoding='utf-8') as f:
        survey = json.load(f)

    description = (
        "Encuesta de intoxicaciones en Pequeños Animales\n"
        "Objetivos\n"
        "Mediante la presente encuesta se pretende recabar información sobre las principales\n"
        "intoxicaciones observadas en pequeños animales atendidos en clínicas veterinarias del\n"
        "Uruguay. La información proporcionada deberá corresponder a casos registrados en el periodo\n"
        "comprendido entre junio de 2025 y la actualidad.\n"
        "Los datos obtenidos serán utilizados en el marco de un proyecto de investigación CIDEC y de\n"
        "una tesis de grado de la Facultad de Veterinaria (UdelaR). A partir de los resultados relevados,\n"
        "se busca contribuir a la puesta a punto en el Laboratorio de Toxicología en dicha institución,\n"
        "de técnicas analíticas que permitan mejorar el diagnóstico de las intoxicaciones mencionadas.\n"
        "La encuesta está organizada en tres bloques: datos generales, principales tóxicos en\n"
        "veterinaria y consideraciones finales. Las preguntas referidas a los principales tóxicos se\n"
        "agrupan en las siguientes categorías: pesticidas, medicamentos de uso veterinario,\n"
        "medicamentos de uso humano, plantas y miscelanea.\n"
        "Completar esta encuesta le llevará solo unos minutos. Toda información proporcionada es\n"
        "estrictamente confidencial y se utilizará exclusivamente con fines académicos y de\n"
        "investigación."
    )

    return render_template('survey.html', survey=survey, sid='new', description=description)

@app.route('/submit/<sid>', methods=['POST'])
def submit(sid):
    payload = request.get_json() or request.form.to_dict(flat=False)
    # normalize form-encoded data
    if isinstance(payload, dict) and all(isinstance(v, list) and len(v) == 1 for v in payload.values()):
        data = {k: v[0] for k, v in payload.items()}
    else:
        data = payload

    survey = load_survey(sid)
    if survey is None:
        return jsonify({'status': 'error', 'errors': ['Encuesta no encontrada']}), 404

    # helper to evaluate showIf conditions based on submitted data
    def eval_showif(cond):
        if not isinstance(cond, dict):
            return False
        for key, val in cond.items():
            v = data.get(key)
            if isinstance(v, list):
                if val not in v:
                    return False
            else:
                if v is None or str(v) != str(val):
                    return False
        return True

    errors = []
    # basic validations: required fields (respectando showIf), option validity, simple types
    numeric_fields = {'age_years', 'weight_kg', 'time_since_exposure_hours'}
    email_field = 'contact_email'

    # Funciones auxiliares para validaciones seguras
    def safe_float(value):
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    def safe_int(value):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None

    for q in survey.get('questions', []):
        name = q.get('name')
        # skip validation if hidden by showIf
        showif = q.get('showIf')
        visible = True
        if showif:
            visible = eval_showif(showif)
        if not visible:
            continue

        if q.get('required'):
            val = data.get(name)
            if val is None or (isinstance(val, str) and val.strip() == ''):
                errors.append(f'El campo "{name}" es obligatorio')
                continue

        # validate option types
        if q.get('type') in ('radio', 'select') and name in data:
            opt = data.get(name)
            if opt not in q.get('options', []):
                errors.append(f'Valor inválido para "{name}"')

        # validation rules from survey JSON
        v = q.get('validation') or {}
        if v:
            # type:number
            if v.get('type') == 'number' and name in data and data.get(name) not in (None, ''):
                try:
                    valnum = float(str(data.get(name)).replace(',', '.'))
                    # Validaciones adicionales para conversiones
                    min_value = safe_float(v.get('min'))
                    if min_value is not None and valnum < min_value:
                        errors.append(f'El campo "{name}" es menor que el mínimo {v.get("min")}')
                    max_value = safe_float(v.get('max'))
                    if max_value is not None and valnum > max_value:
                        errors.append(f'El campo "{name}" es mayor que el máximo {v.get("max")}')
                except Exception:
                    errors.append(f'El campo "{name}" debe ser numérico')

            # pattern
            pattern = v.get('pattern')
            if pattern is not None and isinstance(pattern, str):
                try:
                    if not re.match(pattern, str(data.get(name))):
                        errors.append(f'El campo "{name}" no cumple el patrón requerido')
                except re.error:
                    pass

            # min/max length
            min_length = safe_int(v.get('minLength'))
            if min_length is not None and data.get(name) is not None and len(str(data.get(name))) < min_length:
                errors.append(f'El campo "{name}" debe tener al menos {v.get("minLength")} caracteres')
            max_length = safe_int(v.get('maxLength'))
            if max_length is not None and data.get(name) is not None and len(str(data.get(name))) > max_length:
                errors.append(f'El campo "{name}" debe tener como máximo {v.get("maxLength")} caracteres')

        # fallback checks (kept for backward compatibility)
        if name in numeric_fields and name in data and data.get(name) not in (None, '') and not v.get('type'):
            try:
                float(str(data.get(name)).replace(',', '.'))
            except Exception:
                errors.append(f'El campo "{name}" debe ser numérico')

        # email check (if no pattern provided)
        if name == email_field and data.get(name) and not (v and v.get('pattern')):
            email = data.get(name)
            if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(email)):
                errors.append('Correo electrónico inválido')

    if errors:
        return jsonify({'status': 'error', 'errors': errors}), 400

    # Respect respondent anonymity choice: if respondent selected anonymous, remove PII before storing
    try:
        if str(data.get('respondent_anonymous', '')).upper() == 'SI':
            data.pop('owner_name', None)
            data.pop('contact_email', None)
            # Also ensure allow_contact is set to NO for safety
            data['allow_contact'] = ['NO']
    except Exception:
        pass

    init_db()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('INSERT INTO responses (survey_id, response_json, submitted_at) VALUES (?, ?, ?)',
                (sid, json.dumps(data, ensure_ascii=False), datetime.utcnow().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({'status': 'ok'})

@app.route('/results/<sid>')
@require_admin
def results(sid):
    init_db()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('SELECT id, response_json, submitted_at FROM responses WHERE survey_id = ? ORDER BY id DESC', (sid,))
    rows = cur.fetchall()
    conn.close()
    responses = [{'id': r[0], 'response': json.loads(r[1]), 'submitted_at': r[2]} for r in rows]
    survey = load_survey(sid)
    return render_template('results.html', sid=sid, responses=responses, survey=survey)

@app.route('/results/<sid>/export')
@require_admin
def export_csv(sid):
    init_db()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute('SELECT response_json FROM responses WHERE survey_id = ?', (sid,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        return 'No responses', 404
    # options via query params: ?anonymize=1, ?anonymize_all=1, ?pseudonymize=mysalt
    anonymize = request.args.get('anonymize') in ('1', 'true', 'yes')
    anonymize_all = request.args.get('anonymize_all') in ('1', 'true', 'yes')
    pseudonymize = request.args.get('pseudonymize')

    # collect all responses as dicts
    resp_list = [json.loads(r[0]) for r in rows]
    # determine union of fieldnames
    fieldnames = set()
    for resp in resp_list:
        fieldnames.update(resp.keys())
    fieldnames = list(sorted(fieldnames))

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    def sha256(s):
        return hashlib.sha256(s.encode('utf-8')).hexdigest()

    for resp in resp_list:
        row = dict(resp)
        if anonymize_all or (anonymize and str(row.get('respondent_anonymous','')).upper() == 'SI'):
            row.pop('owner_name', None)
            row.pop('contact_email', None)
            row['allow_contact'] = 'NO'
        elif pseudonymize:
            if 'owner_name' in row and row.get('owner_name'):
                row['owner_name'] = sha256(pseudonymize + str(row['owner_name']))
            if 'contact_email' in row and row.get('contact_email'):
                row['contact_email'] = sha256(pseudonymize + str(row['contact_email']))
        writer.writerow(row)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True,
                     download_name=f'{sid}_responses.csv')

@app.route('/load_structured_survey', methods=['GET'])
def load_structured_survey():
    structured_survey_path = os.path.join(SURVEYS_DIR, 'structured_survey.json')
    if not os.path.exists(structured_survey_path):
        return jsonify({'status': 'error', 'message': 'Structured survey not found'}), 404

    with open(structured_survey_path, 'r', encoding='utf-8') as f:
        structured_survey = json.load(f)

    return jsonify({'status': 'success', 'survey': structured_survey})

@app.route('/api/surveys', methods=['GET'])
def get_survey():
    survey_path = os.path.join(SURVEYS_DIR, 'encuesta_version.json')
    if not os.path.exists(survey_path):
        return jsonify({'status': 'error', 'message': 'Survey not found'}), 404

    with open(survey_path, 'r', encoding='utf-8') as f:
        survey_data = json.load(f)

    # Include `showIf` logic in the survey data
    for question in survey_data.get('questions', []):
        if 'showIf' in question:
            question['showIf'] = question['showIf']  # Ensure `showIf` is included as-is

    return jsonify({'status': 'success', 'survey': survey_data}), 200

def append_to_excel(survey_id, data, survey_questions):
    """Append a response row to the Excel file for this survey."""
    os.makedirs(EXCEL_DIR, exist_ok=True)
    filepath = os.path.join(EXCEL_DIR, f'{survey_id}_responses.xlsx')

    # Build ordered question list for column headers
    q_ids = [q['id'] for q in survey_questions]
    q_texts = {q['id']: q['text'] for q in survey_questions}

    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active or wb.create_sheet('Respuestas')
    else:
        wb = openpyxl.Workbook()
        ws = wb.active or wb.create_sheet('Respuestas')
        ws.title = 'Respuestas'

        # Header row — submission metadata
        header_meta = ['ID envío', 'Fecha y hora']
        header_qs   = [f'{qid}. {q_texts.get(qid, qid)}' for qid in q_ids]
        headers     = header_meta + header_qs

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='6B21A8')
        header_align = Alignment(wrap_text=True, vertical='center')

        ws.append(headers)
        for cell in ws[1]:
            cell.font   = header_font
            cell.fill   = header_fill
            cell.alignment = header_align

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 22
        for col_idx in range(3, len(headers) + 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = 40

        ws.row_dimensions[1].height = 50
        ws.freeze_panes = 'C2'

    # Next submission id = current row count (excluding header)
    submission_id = ws.max_row  # header row counts as 1, so first data row → id=1

    # Build row values
    def fmt(val):
        if isinstance(val, list):
            return ' | '.join(str(v) for v in val)
        if isinstance(val, dict):
            return ' | '.join(f'{k}: {v}' for k, v in val.items())
        return str(val) if val is not None else ''

    row = [submission_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    for qid in q_ids:
        row.append(fmt(data.get(qid, '')))

    ws.append(row)

    # Zebra striping
    row_idx = ws.max_row
    fill = PatternFill('solid', fgColor='F3E8FF') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(filepath)
    return filepath


@app.route('/api/submit', methods=['POST'])
def submit_response():
    data = request.get_json()
    if not data or 'survey_id' not in data or 'responses' not in data:
        return jsonify({'status': 'error', 'message': 'Invalid data'}), 400

    survey_id = data['survey_id']
    responses = data['responses']

    # Load survey for question order / texts
    survey = load_survey(survey_id)
    questions = survey.get('questions', []) if survey else []

    # 1. Save to SQLite (existing)
    init_db()
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    cur.execute(
        'INSERT INTO responses (survey_id, response_json, submitted_at) VALUES (?, ?, ?)',
        (survey_id, json.dumps(responses, ensure_ascii=False), datetime.now().isoformat())
    )
    conn.commit()
    conn.close()

    # 2. Save to PostgreSQL
    save_to_postgres(survey_id, responses)

    # 3. Append to Excel file
    try:
        append_to_excel(survey_id, responses, questions)
    except Exception as e:
        print(f'[Excel] Error: {e}')

    return jsonify({'status': 'success', 'message': 'Respuesta guardada exitosamente'}), 201


@app.route('/api/export/excel/<sid>')
@require_admin
def export_excel(sid):
    filepath = os.path.join(EXCEL_DIR, f'{sid}_responses.xlsx')
    if not os.path.exists(filepath):
        return jsonify({'status': 'error', 'message': 'No hay respuestas aún'}), 404
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{sid}_respuestas.xlsx'
    )

@app.route('/surveys', methods=['GET'])
def get_surveys():
    surveys = []
    if os.path.exists(SURVEYS_DIR):
        for fn in os.listdir(SURVEYS_DIR):
            if fn.endswith('.json'):
                sid = fn[:-5]
                with open(os.path.join(SURVEYS_DIR, fn), 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                surveys.append({'id': sid, 'title': meta.get('title', sid)})
    return jsonify({'status': 'success', 'surveys': surveys}), 200

@app.route('/surveys/main_encuesta.json', methods=['GET'])
def get_main_encuesta():
    survey_path = os.path.join(SURVEYS_DIR, 'main_encuesta.json')
    if not os.path.exists(survey_path):
        return jsonify({'status': 'error', 'message': 'Survey not found'}), 404

    with open(survey_path, 'r', encoding='utf-8') as f:
        survey_data = json.load(f)

    return jsonify({'status': 'success', 'survey': survey_data}), 200

@app.route('/surveys/<sid>.json', methods=['GET'])
def get_survey_json(sid):
    survey_path = os.path.join(SURVEYS_DIR, f'{sid}.json')
    if not os.path.exists(survey_path):
        return jsonify({'status': 'error', 'message': 'Survey not found'}), 404
    with open(survey_path, 'r', encoding='utf-8') as f:
        survey_data = json.load(f)
    return jsonify({'status': 'success', 'survey': survey_data}), 200

# ---------------------------------------------------------------------------
# Serve React SPA (catch-all) — must be the LAST route defined
# Flask API routes above take priority; everything else serves the React app.
# ---------------------------------------------------------------------------
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve_react(path):
    if os.path.isdir(REACT_BUILD_DIR):
        # Serve real static files from the React build (JS, CSS, images)
        target = os.path.join(REACT_BUILD_DIR, path)
        if path and os.path.isfile(target):
            return send_from_directory(REACT_BUILD_DIR, path)
        # All other paths → index.html (React Router handles the rest)
        return send_from_directory(REACT_BUILD_DIR, 'index.html')
    # Fallback: original Jinja2 index when React build is not present
    surveys = []
    if os.path.exists(SURVEYS_DIR):
        for fn in os.listdir(SURVEYS_DIR):
            if fn.endswith('.json'):
                sid = fn[:-5]
                with open(os.path.join(SURVEYS_DIR, fn), 'r', encoding='utf-8') as f:
                    meta = json.load(f)
                surveys.append({'id': sid, 'title': meta.get('title', sid)})
    return render_template('index.html', surveys=surveys)

if __name__ == '__main__':
    os.makedirs(SURVEYS_DIR, exist_ok=True)
    init_db()

    # Argument parser for dynamic port configuration
    parser = argparse.ArgumentParser(description='Run the Flask application.')
    parser.add_argument('--port', type=int, default=5000, help='Port to run the server on (default: 5000)')
    args = parser.parse_args()

    # Auto-open browser when running as packaged exe
    if getattr(sys, 'frozen', False):
        import threading, webbrowser
        url = f'http://127.0.0.1:{args.port}'
        threading.Timer(1.5, lambda: webbrowser.open(url)).start()
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(host='127.0.0.1', port=args.port, debug=debug)
